# -*- coding: utf-8 -*-
"""
发票处理工具 - 自动识别PDF发票，提取信息，生成报销清单并重命名文件
"""

import os
import re
import sys
import threading
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pdfplumber
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

APP_NAME = "发票处理工具"
APP_VERSION = "1.0"
WINDOW_SIZE = "1100x680"

# Treeview column definitions
COLUMNS = [
    #  (id,       header,     width, anchor)
    ("seq",       "序号",      50,   "center"),
    ("filename",  "文件名",   200,   "w"),
    ("item_name", "项目内容", 250,   "w"),
    ("date",      "开票日期", 130,   "center"),
    ("total",     "价税合计",  90,   "e"),
    ("inv_num",   "发票号码", 220,   "w"),
]


# ================================================================
#  PDF 提取逻辑
# ================================================================

def extract_invoice_number(text):
    """提取发票号码"""
    # 方式1：行内格式  发票号码：XXXXX / 发 票 号 码 ：XXXXX
    for line in text.split('\n'):
        m = re.search(r'发\s*票\s*号\s*码\s*[：:]\s*(\d{10,})', line)
        if m:
            return m.group(1)
    # 方式2：监制章格式  制 XXXXX
    for line in text.split('\n'):
        m = re.search(r'制\s*(\d{20,})', line)
        if m:
            return m.group(1)
    return ''


def extract_date(text):
    """提取开票日期"""
    for line in text.split('\n'):
        m = re.search(
            r'开\s*票\s*日\s*期\s*[：:]\s*'
            r'(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)', line)
        if m:
            return re.sub(r'\s+', '', m.group(1))
        # 税务局行（监制章格式）
        if re.search(r'税务', line):
            m2 = re.search(r'(\d{4}\s*年\s*\d{1,2}\s*月\s*\d{1,2}\s*日)', line)
            if m2:
                return re.sub(r'\s+', '', m2.group(1))
    return ''


def extract_total(text):
    """提取价税合计（小写金额）"""
    m = re.search(r'[（(]\s*小\s*写\s*[）)]\s*¥\s*([\d,]+\.?\d*)', text)
    if m:
        return float(m.group(1).replace(',', ''))
    return None


def extract_item_name(text):
    """提取第一个项目名称"""
    # 预处理：合并 PDF 跨行中文（如 "平\n键销" → "平键销"）
    clean = re.sub(r'(?<=[一-鿿])\n\s*(?=[一-鿿])', '', text)
    clean = re.sub(r' {2,}', ' ', clean)

    cat_match = re.search(r'\*[^*\n]+\*', clean)
    if not cat_match:
        return ''

    section = clean[cat_match.end():]
    heji = re.search(r'合\s*计', section)
    if heji:
        section = section[:heji.start()]

    # 取到首个「单位 + 数量」之前
    unit_pat = (
        r'(?:PCS|件|个|台|套|只|米|条|把|盒|支|根|块|张|本|份'
        r'|袋|桶|组|瓶|包|罐|对|双)\s+\d'
    )
    um = re.search(unit_pat, section)
    if um:
        name = section[:um.start()].strip()
    else:
        first_line = section.split('\n')[0]
        m = re.search(r'[^\n\d]{2,}', first_line)
        name = m.group(0).strip() if m else first_line.strip()

    # 检查下一行是否为名称续行（纯中文、无数字、无 * 类别标记）
    if um:
        after_unit = section[um.end():]
        next_nl = after_unit.find('\n')
        if next_nl >= 0:
            next_line = after_unit[next_nl + 1:].strip().split('\n')[0].strip()
            if (next_line
                    and not re.search(r'\d', next_line)
                    and not re.search(r'\*', next_line)
                    and re.match(r'^[一-鿿]+$', next_line)):
                name = name + next_line

    name = re.sub(r'\s*[*]\s*$', '', name.strip())
    return re.sub(r'\s+', ' ', name)


def extract_pdf(filepath):
    """从单个 PDF 提取发票信息，返回 dict"""
    pdf = pdfplumber.open(filepath)
    text = ''
    for page in pdf.pages:
        text += (page.extract_text() or '') + '\n'
    pdf.close()

    inv_num = extract_invoice_number(text)
    date = extract_date(text)
    total = extract_total(text)
    item_name = extract_item_name(text)

    if not item_name:
        item_name = re.sub(r'\d+元$', '',
                           os.path.basename(filepath).replace('.pdf', ''))

    return {
        'filename': os.path.basename(filepath),
        'item_name': item_name,
        'date': date,
        'total': total,
        'inv_num': inv_num,
    }


# ================================================================
#  Excel 写入
# ================================================================

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin'),
)

HEADERS = {
    1: '序号', 2: '项目内容', 3: '时间', 4: '费用',
    5: '报销凭证', 6: '备注', 7: '报销人', 8: '学号',
    9: '合计', 10: '备注',
}


def write_excel(filepath, data_list, template_path=None):
    """
    将发票数据写入 Excel。
    data_list: [{'seq', 'item_name', 'date', 'total', 'inv_num'}, ...]
    """
    if template_path and os.path.exists(template_path):
        wb = openpyxl.load_workbook(template_path)
        ws = wb.active
        # 取消合并单元格
        for mr in list(ws.merged_cells.ranges):
            ws.unmerge_cells(str(mr))
        # 清除数据行
        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(row=r, column=c).value = None
    else:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Sheet1'
        header_font = Font(bold=True)
        header_fill = PatternFill('solid', fgColor='D9E1F2')
        for col, title in HEADERS.items():
            cell = ws.cell(row=1, column=col, value=title)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = THIN_BORDER
        # 列宽
        widths = {1: 6, 2: 35, 3: 16, 4: 12, 5: 12,
                  6: 30, 7: 10, 8: 15, 9: 10, 10: 10}
        for c, w in widths.items():
            ws.column_dimensions[openpyxl.utils.get_column_letter(c)].width = w

    # 写入数据
    for i, d in enumerate(data_list):
        row = i + 2
        ws.cell(row=row, column=1, value=d['seq'])
        ws.cell(row=row, column=2, value=d['item_name'])
        ws.cell(row=row, column=3, value=d['date'])
        ws.cell(row=row, column=4, value=d['total'])
        ws.cell(row=row, column=5, value='电子发票')
        ws.cell(row=row, column=6, value=f"发票号码:{d['inv_num']}")

    last = len(data_list) + 1
    if last >= 2:
        ws.cell(row=2, column=9, value=f'=SUM(D2:D{last})')

    wb.save(filepath)


# ================================================================
#  GUI 应用
# ================================================================

class InvoiceApp:
    def __init__(self):
        self.root = tk.Tk()
        self.root.title(f"{APP_NAME} v{APP_VERSION}")
        self.root.geometry(WINDOW_SIZE)
        self.root.minsize(900, 500)

        self.data = []           # list[dict]
        self.folder = ''
        self.template = None
        self._edit_widget = None

        self._apply_style()
        self._build_ui()

    # ---- 样式 ----
    def _apply_style(self):
        style = ttk.Style()
        try:
            style.theme_use('vista')
        except tk.TclError:
            try:
                style.theme_use('clam')
            except tk.TclError:
                pass

        style.configure("TButton", padding=5)
        style.configure("Title.TLabel", font=("微软雅黑", 14, "bold"))
        style.configure("Summary.TLabel", font=("微软雅黑", 10))
        style.configure("Status.TLabel", font=("微软雅黑", 9))

        style.configure(
            "Invoice.Treeview",
            font=("微软雅黑", 9),
            rowheight=26,
            borderwidth=1,
        )
        style.configure(
            "Invoice.Treeview.Heading",
            font=("微软雅黑", 9, "bold"),
        )
        style.map("Invoice.Treeview",
                   background=[("selected", "#3471CE")],
                   foreground=[("selected", "#FFFFFF")])

    # ---- 界面构建 ----
    def _build_ui(self):
        pad = {"padx": 10, "pady": 5}

        # --- 顶部：文件夹选择 ---
        top = ttk.Frame(self.root, padding=(10, 8))
        top.pack(fill=tk.X)

        ttk.Label(top, text="发票文件夹:").pack(side=tk.LEFT)
        self.path_var = tk.StringVar()
        self.path_entry = ttk.Entry(top, textvariable=self.path_var, width=55)
        self.path_entry.pack(side=tk.LEFT, padx=(5, 5), fill=tk.X, expand=True)

        ttk.Button(top, text="浏览", command=self._browse, width=7).pack(side=tk.LEFT, padx=2)
        self.scan_btn = ttk.Button(top, text="扫描", command=self._scan, width=7)
        self.scan_btn.pack(side=tk.LEFT)

        # --- 中部：表格 ---
        tree_frame = ttk.Frame(self.root, padding=(10, 2))
        tree_frame.pack(fill=tk.BOTH, expand=True)

        self.tree = ttk.Treeview(
            tree_frame,
            columns=[c[0] for c in COLUMNS],
            show="headings",
            style="Invoice.Treeview",
            selectmode="browse",
        )
        for cid, header, width, anchor in COLUMNS:
            self.tree.heading(cid, text=header, anchor="center")
            self.tree.column(cid, width=width, minwidth=40, anchor=anchor)

        vsb = ttk.Scrollbar(tree_frame, orient=tk.VERTICAL, command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient=tk.HORIZONTAL, command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        tree_frame.grid_rowconfigure(0, weight=1)
        tree_frame.grid_columnconfigure(0, weight=1)

        # 斑马纹
        self.tree.tag_configure("even", background="#F0F4F8")
        self.tree.tag_configure("odd", background="#FFFFFF")

        # 双击编辑
        self.tree.bind("<Double-1>", self._on_dblclick)

        # 提示文字
        self.hint_label = ttk.Label(tree_frame, text="双击单元格可编辑",
                                    foreground="gray")
        self.hint_label.grid(row=2, column=0, pady=2, sticky="w")

        # --- 底部：统计 + 导出按钮 ---
        bottom = ttk.Frame(self.root, padding=(10, 5))
        bottom.pack(fill=tk.X)

        self.summary_var = tk.StringVar(value="")
        ttk.Label(bottom, textvariable=self.summary_var,
                  style="Summary.TLabel").pack(side=tk.LEFT)

        self.export_btn = ttk.Button(
            bottom, text="导出 Excel 并重命名",
            command=self._export, state=tk.DISABLED,
        )
        self.export_btn.pack(side=tk.RIGHT, padx=(10, 0))

        sep = ttk.Separator(self.root, orient=tk.HORIZONTAL)
        sep.pack(fill=tk.X, padx=10)

        status_frame = ttk.Frame(self.root, padding=(10, 3))
        status_frame.pack(fill=tk.X, side=tk.BOTTOM)
        self.status_var = tk.StringVar(value="就绪")
        ttk.Label(status_frame, textvariable=self.status_var,
                  style="Status.TLabel", foreground="gray").pack(side=tk.LEFT)

    # ---- 浏览文件夹 ----
    def _browse(self):
        folder = filedialog.askdirectory(title="选择发票所在文件夹")
        if folder:
            self.path_var.set(folder)
            self._scan()

    # ---- 扫描 ----
    def _scan(self):
        folder = self.path_var.get().strip()
        if not folder or not os.path.isdir(folder):
            messagebox.showwarning("提示", "请先选择有效的文件夹")
            return

        self.folder = folder
        self._find_template()

        # 筛选新发票 PDF
        pdfs = sorted([
            f for f in os.listdir(folder)
            if f.lower().endswith('.pdf')
            and '报价单' not in f
            and not re.match(r'^\d+\s', f)
        ])

        if not pdfs:
            messagebox.showinfo("提示", "该文件夹下没有新的发票 PDF 文件\n（已有序号前缀的文件将被跳过）")
            self.export_btn.config(state=tk.DISABLED)
            self.summary_var.set("")
            return

        # 禁用按钮、显示进度
        self.scan_btn.config(state=tk.DISABLED)
        self.export_btn.config(state=tk.DISABLED)
        self.status_var.set(f"正在扫描 {len(pdfs)} 个文件…")

        # 进度窗口
        prog = tk.Toplevel(self.root)
        prog.title("扫描中")
        prog.geometry("460x110")
        prog.resizable(False, False)
        prog.transient(self.root)
        prog.grab_set()
        ttk.Label(prog, text="正在扫描发票，请稍候…",
                  font=("微软雅黑", 10)).pack(pady=(15, 8))
        pbar = ttk.Progressbar(prog, length=400, maximum=len(pdfs), mode="determinate")
        pbar.pack(padx=20)
        plabel = ttk.Label(prog, text="", foreground="gray")
        plabel.pack(pady=(5, 0))

        # 在后台线程中处理，避免阻塞 UI
        def worker():
            results = []
            for i, fname in enumerate(pdfs):
                fpath = os.path.join(folder, fname)
                plabel.config(text=f"[{i+1}/{len(pdfs)}] {fname}")
                pbar['value'] = i + 1
                prog.update_idletasks()
                try:
                    info = extract_pdf(fpath)
                except Exception as e:
                    info = {
                        'filename': fname,
                        'item_name': f'提取失败: {e}',
                        'date': '', 'total': None, 'inv_num': '',
                    }
                results.append(info)

            prog.destroy()
            self._on_scan_done(results)

        t = threading.Thread(target=worker, daemon=True)
        t.start()

    def _on_scan_done(self, results):
        """扫描完成回调（在主线程中执行）"""
        self.scan_btn.config(state=tk.NORMAL)

        # 分配序号
        next_seq = self._next_seq()
        for item in results:
            item['seq'] = next_seq
            next_seq += 1

        self.data = results
        self._populate_table()

        total_sum = sum(d['total'] for d in self.data if d['total'] is not None)
        self.summary_var.set(
            f"共 {len(self.data)} 张发票    "
            f"合计金额: ¥{total_sum:,.2f}"
        )
        self.status_var.set(
            f"扫描完成，找到 {len(self.data)} 张发票（起始序号 "
            f"{self.data[0]['seq']}）"
        )
        self.export_btn.config(state=tk.NORMAL)

    def _populate_table(self):
        self.tree.delete(*self.tree.get_children())
        for i, d in enumerate(self.data):
            tag = "even" if i % 2 == 0 else "odd"
            total_str = f"¥{d['total']:,.2f}" if d['total'] is not None else ""
            self.tree.insert(
                "", "end",
                values=(d['seq'], d['filename'], d['item_name'],
                        d['date'], total_str, d['inv_num']),
                tags=(tag,),
            )

    def _next_seq(self):
        """确定下一个可用序号"""
        max_seq = 0
        if self.folder:
            for f in os.listdir(self.folder):
                m = re.match(r'^(\d+)\s', f)
                if m:
                    max_seq = max(max_seq, int(m.group(1)))

        if self.template and os.path.exists(self.template):
            try:
                wb = openpyxl.load_workbook(self.template)
                ws = wb.active
                for r in range(2, ws.max_row + 1):
                    v = ws.cell(row=r, column=1).value
                    if isinstance(v, (int, float)):
                        max_seq = max(max_seq, int(v))
            except Exception:
                pass
        return max_seq + 1

    def _find_template(self):
        """在文件夹中查找模板文件"""
        self.template = None
        if not self.folder:
            return
        for f in os.listdir(self.folder):
            if f.endswith('.xlsx') and '模板' in f:
                self.template = os.path.join(self.folder, f)
                return

    # ---- 单元格编辑 ----
    def _on_dblclick(self, event):
        if self._edit_widget:
            self._edit_widget.destroy()
            self._edit_widget = None

        region = self.tree.identify("region", event.x, event.y)
        if region != "cell":
            return

        col_id = self.tree.identify_column(event.x)   # #1, #2, ...
        col_idx = int(col_id.replace('#', '')) - 1
        item_id = self.tree.identify_row(event.y)
        if not item_id:
            return

        keys = [c[0] for c in COLUMNS]
        key = keys[col_idx]

        # 序号列不允许编辑
        if key == 'seq':
            return

        iid = self.tree.index(item_id)
        current = self.data[iid].get(key, '')
        if current is None:
            current = ''
        if key == 'total' and isinstance(current, (int, float)):
            current = f"{current:.2f}"

        x, y, w, h = self.tree.bbox(item_id, col_id)

        entry = tk.Entry(self.tree, font=("微软雅黑", 9))
        entry.place(x=x, y=y, width=w, height=h)
        entry.insert(0, str(current))
        entry.select_range(0, tk.END)
        entry.focus_set()
        self._edit_widget = entry

        def save(evt=None):
            val = entry.get()
            if key == 'total':
                try:
                    val = float(val.replace(',', '').replace('¥', ''))
                    self.data[iid]['total'] = val
                    display = f"¥{val:,.2f}"
                except ValueError:
                    self.data[iid]['total'] = None
                    display = ""
            else:
                self.data[iid][key] = val
                display = val

            vals = list(self.tree.item(item_id, 'values'))
            vals[col_idx] = display
            self.tree.item(item_id, values=vals)

            # 更新合计
            total_sum = sum(d['total'] for d in self.data if d['total'] is not None)
            self.summary_var.set(
                f"共 {len(self.data)} 张发票    "
                f"合计金额: ¥{total_sum:,.2f}"
            )

            entry.destroy()
            self._edit_widget = None

        def cancel(evt=None):
            entry.destroy()
            self._edit_widget = None

        entry.bind("<Return>", save)
        entry.bind("<Escape>", cancel)
        entry.bind("<FocusOut>", save)

    # ---- 导出 ----
    def _export(self):
        if not self.data:
            messagebox.showwarning("提示", "没有可导出的数据")
            return

        # 确认对话框
        total_sum = sum(d['total'] for d in self.data if d['total'] is not None)
        msg = (
            f"即将导出 {len(self.data)} 张发票信息并重命名 PDF 文件。\n\n"
            f"合计金额: ¥{total_sum:,.2f}\n"
            f"序号范围: {self.data[0]['seq']} ~ {self.data[-1]['seq']}\n"
        )
        if self.template:
            msg += f"\n使用模板: {os.path.basename(self.template)}"
        else:
            msg += "\n（未找到模板文件，将新建 Excel）"

        if not messagebox.askyesno("确认导出", msg):
            return

        # 保存对话框
        default_name = "报销清单.xlsx"
        if self.template:
            # 优先保存到模板同名文件
            default_name = os.path.basename(self.template).replace('模板', '').strip('-')
            if not default_name:
                default_name = "报销清单.xlsx"

        filepath = filedialog.asksaveasfilename(
            title="保存报销清单",
            defaultextension=".xlsx",
            initialdir=self.folder,
            initialfile=default_name,
            filetypes=[("Excel 文件", "*.xlsx")],
        )
        if not filepath:
            return

        # 写入 Excel
        try:
            excel_data = [{
                'seq': d['seq'],
                'item_name': d['item_name'],
                'date': d['date'],
                'total': d['total'],
                'inv_num': d['inv_num'],
            } for d in self.data]
            write_excel(filepath, excel_data, self.template)
        except PermissionError:
            messagebox.showerror("错误",
                                 "无法写入文件，请关闭 Excel 后重试。")
            return
        except Exception as e:
            messagebox.showerror("错误", f"写入 Excel 失败:\n{e}")
            return

        # 重命名 PDF
        renamed = 0
        rename_errors = []
        for d in self.data:
            old = os.path.join(self.folder, d['filename'])
            new_name = f"{d['seq']} {d['filename']}"
            new = os.path.join(self.folder, new_name)
            if os.path.exists(old) and old != new:
                try:
                    os.rename(old, new)
                    d['filename'] = new_name
                    renamed += 1
                except Exception as e:
                    rename_errors.append(f"{d['filename']}: {e}")

        self._populate_table()
        self.export_btn.config(state=tk.DISABLED)

        # 完成提示
        result_msg = (
            f"导出完成！\n\n"
            f"  Excel 写入: {len(self.data)} 条记录\n"
            f"  合计金额: ¥{total_sum:,.2f}\n"
            f"  PDF 重命名: {renamed} 个\n"
            f"  保存位置: {os.path.basename(filepath)}"
        )
        if rename_errors:
            result_msg += f"\n\n重命名失败:\n" + "\n".join(rename_errors)

        messagebox.showinfo("完成", result_msg)
        self.status_var.set(
            f"已导出: {len(self.data)} 条记录, "
            f"¥{total_sum:,.2f}, "
            f"重命名 {renamed} 个文件"
        )

    # ---- 运行 ----
    def run(self):
        self.root.mainloop()


# ================================================================
#  入口
# ================================================================

if __name__ == "__main__":
    # 支持直接拖拽文件夹到图标启动
    if len(sys.argv) > 1 and os.path.isdir(sys.argv[1]):
        app = InvoiceApp()
        app.path_var.set(sys.argv[1])
        app.root.after(200, app._scan)
        app.run()
    else:
        app = InvoiceApp()
        app.run()
